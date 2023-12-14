from pathlib import Path
import math
import json
import pandas as pd
import re
import os
import sys
from scipy.spatial import ConvexHull
import numpy as np
import openpyxl
import subprocess

from . import HISTORY_PATH, TREATMENT_REP_FORMAT, DAY_FORMAT

import logging
logger = logging.getLogger(__name__)


def initiator():
    # check if HISTORY_PATH Exists
    if not HISTORY_PATH.exists():
        # Create the directory
        HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)

        #make an empty .json history file
        with open(HISTORY_PATH, "w") as f:
            json.dump({}, f, indent=4)

        logger.info("No history file found! Creating new history file...")


def num_to_ord(input_number):
    suf = lambda n: "%d%s"%(n,{1:"st",2:"nd",3:"rd"}.get(n%100 if (n%100)<20 else n%10,"th"))
    if isinstance(input_number, list):
        return [suf(i) for i in input_number]
    elif isinstance(input_number, int):
        return suf(input_number)
    else:
        raise TypeError("Input must be an integer or a list of integers")
    
def ord_to_num(input_ord):
    if isinstance(input_ord, list):
        return [int(re.findall(r'\d+', i)[0]) for i in input_ord]
    elif isinstance(input_ord, str):
        return int(re.findall(r'\d+', input_ord)[0])
    else:
        raise TypeError("Input must be a string or a list of strings")
    
def index_to_char(input_number): # turn 0 to A
    if isinstance(input_number, list):
        return [chr(i+65) for i in input_number]
    elif isinstance(input_number, int):
        return chr(input_number+65)
    else:
        raise TypeError("Input must be an integer or a list of integers")
    
def char_to_index(input_char): # turn A to 0
    if isinstance(input_char, list):
        return [ord(i)-65 for i in input_char]
    elif isinstance(input_char, str):
        return ord(input_char)-65

def count_csv_file(directory_path):
    directory = Path(directory_path)
    csv_files = list(directory.glob('*.csv'))

    if len(csv_files) > 0:
        return len(csv_files)
    else:
        return 0
    
def find_the_whole_num(given_string, end_index):

    num_in_string = ""

    i = end_index

    while True:
        num_in_string = given_string[i] + num_in_string
        try:
            previous_char = given_string[i - 1]
        except IndexError:
            break

        if previous_char.isdigit():
            i -= 1
        else:
            break
            
    return int(num_in_string)

def find_batch_num(given_string):
    """
        find "st", "nd", "rd", "th" in a string
        if found, check if the previous character is a number
        if yes, return the number
        if no, return None
    """

    # find all occurrences of "st", "nd", "rd", "th"
    occurrences = [m.start() for m in re.finditer('st|nd|rd|th', given_string)]

    # find the occurrences that have a number before it
    num_occurrence = [occ - 1 for occ in occurrences if given_string[occ - 1].isdigit()]

    if len(num_occurrence) != 1:
        return find_day_num(given_string)
    
    num_occurrence = num_occurrence[0]

    return find_the_whole_num(given_string, num_occurrence)

def find_day_num(given_string):

    day_num = given_string.split(" ")[-1]

    if day_num.isdigit():
        return int(day_num)
    else:
        return None

def find_treatment_num(given_string):

    indicator = given_string.split("-")[0].strip()

    # Check if indicator is number of char
    try:
        indicator = int(indicator)
    except ValueError:
        # change A -> 1, B -> 2, C -> 3, etc.
        try:
            indicator = char_to_index(indicator)
        except Exception as e:
            message = f"[STRUCTURE ERROR] The indicator before the dash char ( - ) in Treatment folder name is unusual!\n{e}"
            logger.error(f"{indicator=}")
            raise Exception(message)

    return indicator
    
# Create event_dict from event binary list
def event_extractor(binary_list, positive_token = None):

    # find unique values in the binary list
    # binary_list is a list
    unique_values = list(set(binary_list))

    if len(unique_values) == 2:
        if positive_token is None:
            # positive_token = non zero value in unique_values
            positive_token = [i for i in unique_values if i != 0][0]
        else:
            if positive_token not in unique_values:
                raise ValueError("The specified positive token is not in the binary list.")
    elif len(unique_values) > 2:
        if positive_token is None:
            raise ValueError("The binary list has more than two unique values. Please specify the positive token.")
        else:
            if positive_token not in unique_values:
                raise ValueError("The specified positive token is not in the binary list.")
    
    binary_list = [1 if i == positive_token else 0 for i in binary_list]

    result = {}
    start, end = None, None

    for i in range(len(binary_list)):
        if binary_list[i] == 1:
            if start is None:
                start = i
            end = i
        elif start is not None:
            result[(start, end)] = end - start + 1
            start, end = None, None

    if start is not None:
        result[(start, end)] = end - start + 1

    return result


# Excel related functions
def check_sheet_existence(file_path, sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if sheet_name in workbook.sheetnames:
        return True
    else:
        return False
    
def remove_sheet_by_name(file_path, sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
        workbook.save(filename=file_path)
        return True
    else:
        return False
    
def append_df_to_excel(filename, df, sheet_name='Sheet1', startcol=None, startrow=None, col_sep = 0, row_sep = 0,
                       truncate_sheet=False, DISPLAY = False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        try:
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startcol=startcol if startcol is not None else 0, 
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            logger.info(f"Successful write to {filename}, sheet={sheet_name} at column = {startcol}, row = {startrow}")
        except Exception as e:
            logger.warning(f"EXCEPTION: {e}")
            logger.warning(f"UNSUCCESS write to {filename}, sheet={sheet_name} at column = {startcol}, row = {startrow}")
            logger.debug(f"df: {df}")

        # wb = openpyxl.load_workbook(filename)
        # ws = wb[sheet_name]
        # row_0 = ws[1]
        return
    

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.workbook = openpyxl.load_workbook(filename)

    # get the last col in the existing Excel sheet
    # if it was not specified explicitly
    if startcol is None and sheet_name in writer.workbook.sheetnames:
        startcol = writer.workbook[sheet_name].max_column + col_sep

    if startrow is None and sheet_name in writer.workbook.sheetnames:
        startrow = writer.workbook[sheet_name].max_row + row_sep
    
    if startcol is None:
        startcol = 0

    if startrow is None:
        startrow = 0
    
    # row_0 = writer.workbook[sheet_name][1]
    # logger.debug(f"Header: {row_0}")
    
    # remove df headers if they exist
    if startrow != 0:
        # take the first row
        first_row = df.iloc[0].astype(str)
        # check if any cell in the first row contains a letter
        has_letter = first_row.str.contains('[a-zA-Z]').any()
        if has_letter:
            df = df.iloc[1:, :]

    # write the dataframe to the existing sheet
    try:
        df.to_excel(writer, sheet_name, startcol=startcol, startrow=startrow, **to_excel_kwargs)
        logger.info(f"Successful write to {filename}/{sheet_name} at column = {startcol}, row = {startrow}")
    except:
        logger.warning(f"UNSUCCESS write to {filename}/{sheet_name} at column = {startcol}, row = {startrow}")

    # close workbook
    writer.close()


def merge_cells(file_path, input_sheet_name = None, input_column_name = 'Shoaling Area', cell_step=3, inplace = True):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if input_sheet_name == None:
        sheet_names = workbook.worksheets
    elif isinstance(input_sheet_name, list):
        sheet_names = input_sheet_name
    elif isinstance(input_sheet_name, str):
        sheet_names = [input_sheet_name]
    elif isinstance(input_sheet_name, int):
        sheet_names = [workbook.worksheets[input_sheet_name]]

    if isinstance(input_column_name, str):
        column_names = [input_column_name]
    elif isinstance(input_column_name, list):
        column_names = input_column_name

    logger.debug(f"Merging {cell_step} rows of {column_names} in {sheet_names}")

    for worksheet in workbook.worksheets:
        if worksheet not in sheet_names:
            continue
        # Find the column index for the "Shoaling Area" header
        for colume_name in column_names:
            shoaling_area_col = None
            for col_idx in range(1, worksheet.max_column+1):
                header = worksheet.cell(row=1, column=col_idx).value
                if header and colume_name in header:
                    shoaling_area_col = col_idx
                    break

            if shoaling_area_col is None:
                print("Column not found.")
            else:
                # Merge every next 3 rows of the Shoaling Area column
                for row_idx in range(2, worksheet.max_row+1, cell_step):
                    value = worksheet.cell(row=row_idx, column=shoaling_area_col).value
                    # print(value)
                    if value is not None:
                        # Merge the current row with the next 2 rows
                        worksheet.merge_cells(start_row=row_idx, start_column=shoaling_area_col, end_row=row_idx+2, end_column=shoaling_area_col)
                    
                    # align the merged cell, horizontal and vertical center
                    worksheet.cell(row=row_idx, column=shoaling_area_col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
    # define output_path
    if inplace == False:
        output_path = file_path[:-5] + '_merged.xlsx'        
    else:
        output_path = file_path    

    # Save the modified workbook
    try:
        workbook.save(filename=output_path)
        logger.debug(f"Merged {cell_step} rows of {column_names} in {sheet_names}")
    except:
        logger.warning(f"UNSUCCESS merge for {file_path}")


def excel_polish(file_path, batch_num=1, inplace=True):

    logger.debug("Polishing excel file...")

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)


    # Adjust the column widths
    # Loop through each sheet in the workbook
    for sheet_name in workbook.sheetnames:

        logger.debug(f"In sheet name: {sheet_name}")

        if "analysis" in sheet_name.lower():
            continue
        # Select the sheet
        sheet = workbook[sheet_name]
        
        # Loop through each column in the sheet
        for col in sheet.columns:
            # Set the width of the column to 17.00 (160 pixels)
            sheet.column_dimensions[col[0].column_letter].width = 17.00

            logger.debug(f"Set column {col[0].column_letter} width to 17.00")
        
        # Enable text wrapping for the header row
        for cell in sheet[1]:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    # Save the modified workbook
    try:
        workbook.save(filename=file_path)
        logger.info(f"Polished completely for {file_path}")
    except:
        logger.info(f"UNSUCCESS polish for {file_path}")

# Others
def open_explorer(path):
    # Check if the given path exists
    if os.path.exists(path):
        real_path = os.path.realpath(path)
        try:
            if sys.platform.startswith('win'):
                # Windows
                subprocess.run(['explorer', real_path])
            elif sys.platform.startswith('darwin'):
                # macOS
                subprocess.run(['open', real_path])
            elif sys.platform.startswith('linux'):
                # Linux
                subprocess.run(['xdg-open', real_path])
            else:
                print(f"Unsupported OS: {sys.platform}")
        except Exception as e:
            print(f"An error occurred: {e}")
    else:
        print("The provided path does not exist.")


def string_diff(string1, string2):
    # Initialize a variable to keep track of the difference count
    diff_count = 0

    # Replace all " " with ""
    str1 = string1.replace(" ", "")
    str2 = string2.replace(" ", "")

    min_len = min(len(str1), len(str2))

    for i in range(min_len):
        if str1[i] != str2[i]:
            diff_count += 1

    diff_count += abs(len(str1) - len(str2))

    return diff_count

def get_folder_info(project_dir, day_num, treatment_char):
    # open HISTORY_PATH
    with open(HISTORY_PATH, 'r') as file:
        data = json.load(file)

    folder_info = []

    print(f"Searching for {project_dir} in {HISTORY_PATH}")

    # check if any project has "DIRECTORY" == project_dir
    for project in data.values():
        FOUND = False
        if project["DIRECTORY"] == str(project_dir):
            FOUND = True
            day_nums = [int(day_name.split(" ")[1]) for day_name in project.keys() if day_name != "DIRECTORY"]
            if day_num in day_nums:
                for treatment_rep in project[DAY_FORMAT.format(day_num)]:
                    if TREATMENT_REP_FORMAT.format(treatment_char) == treatment_rep:
                        folder_info = project[DAY_FORMAT.format(day_num)][treatment_rep][:-1]
                        break
                break
            else:
                logger.error(f"Day {day_num} not found in {project_dir}")
                raise Exception(f"Day {day_num} not found in {project_dir}")
            
    if FOUND == False:
        logger.error(f"Project {project_dir} not found in {HISTORY_PATH}")
        raise Exception(f"Project {project_dir} not found in {HISTORY_PATH}")
    
    return folder_info


def to_int_or_float(given_input):
    output = float(given_input)
        
    if output.is_integer():
        output = int(output)
    
    return output


def find_uncommon_substrings(list_of_strings):
    """
    Extracts the different parts (uncommon substrings) from a list of strings.

    :param list_of_strings: List of strings to compare
    :return: List of uncommon substrings
    """
    if not list_of_strings:
        return []

    # Find the shortest string to avoid index out of range errors
    min_length = min(len(s) for s in list_of_strings)

    # Initialize variables
    differing_indices = []
    prev_diff = False

    # Identify indices where strings differ
    for i in range(min_length):
        chars = {s[i] for s in list_of_strings}
        if len(chars) > 1:
            if not prev_diff:
                differing_indices.append(i)
            prev_diff = True
        else:
            if prev_diff:
                differing_indices.append(i)
            prev_diff = False

    # Extract the uncommon substrings
    uncommon_substrings = set()
    for i in range(0, len(differing_indices), 2):
        start = differing_indices[i]
        end = differing_indices[i + 1]
        substrings = {s[start:end] for s in list_of_strings}
        uncommon_substrings.update(substrings)

    return_list = list(uncommon_substrings)
    return_list = [int(i) if i.isdigit() else i for i in return_list]
    return_list.sort()

    return return_list


def find_uncommon_substrings_in_paths(list_of_paths):
    """
    Extracts the different parts (uncommon substrings) from a list of pathlib.Path objects
    and returns a dictionary mapping each uncommon substring to the corresponding Path object.

    :param list_of_paths: List of pathlib.Path objects to compare
    :return: Dictionary of uncommon substrings and their corresponding Path objects
    """
    if not list_of_paths:
        return {}

    # Extract the name part of each path
    list_of_strings = [p.name for p in list_of_paths]

    # Find the shortest string to avoid index out of range errors
    min_length = min(len(s) for s in list_of_strings)

    # Initialize variables
    differing_indices = []
    prev_diff = False

    # Identify indices where strings differ
    for i in range(min_length):
        chars = {s[i] for s in list_of_strings}
        if len(chars) > 1:
            if not prev_diff:
                differing_indices.append(i)
            prev_diff = True
        else:
            if prev_diff:
                differing_indices.append(i)
            prev_diff = False

    # Extract the uncommon substrings and map them to paths
    uncommon_substrings_dict = {}
    for i in range(0, len(differing_indices), 2):
        start = differing_indices[i]
        end = differing_indices[i + 1] if i + 1 < len(differing_indices) else min_length
        for path in list_of_paths:
            uncommon_substring = path.name[start:end]
            if uncommon_substring:
                uncommon_substrings_dict[uncommon_substring] = path

    return uncommon_substrings_dict


def find_uncommon_substrings_in_paths_general(list_of_paths):
    """
    Extracts the different parts (uncommon substrings) from a list of pathlib.Path objects
    and returns a dictionary mapping each uncommon substring to the corresponding Path object.
    This version focuses on numeric parts as the uncommon substrings.

    :param list_of_paths: List of pathlib.Path objects to compare
    :return: Dictionary of uncommon substrings and their corresponding Path objects
    """
    if not list_of_paths:
        return {}

    # Extract the name part of each path and find all numeric substrings
    substrings = {}
    for path in list_of_paths:
        # Find all numeric substrings in the filename
        numeric_parts = re.findall(r'\d+', path.name)

        # Map each numeric part to the path (assuming each part is unique across the paths)
        for part in numeric_parts:
            substrings[part] = path

    return substrings